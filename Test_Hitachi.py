import openpyxl
import netmiko


def get_nodedata():
    
    workbook = openpyxl.load_workbook('myresource/Topo.xlsx')

    topo = input('\n\n\nPlease enter topology you want to configure : \n\n')
    sheet = workbook['Topo '+topo]

    for i in range(sheet.max_row):
        for o in range(sheet.max_column):
            if(sheet.cell(row=i+2,column=o+1).value):
                print(sheet.cell(row=i+2,column=o+1).value)
    

def main():
    
    while(True):
        get_nodedata()
        break


main()

    


